import pandas as pd
from datetime import datetime
from openpyxl import Workbook


class EmployeeXLSXGenerator:

    def __init__(self, csv_file_name, xlsx_file_name):
        self.csv_file_name = csv_file_name
        self.xlsx_file_name = xlsx_file_name

    def calculate_age(self, dob):
        """Calculate the age based on the date of birth."""
        today = datetime.today()
        age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
        return age

    def generate_xlsx(self):
        try:
            # Load CSV data into a pandas DataFrame
            data = pd.read_csv(self.csv_file_name, encoding='utf-8-sig', sep=';', parse_dates=['Дата народження'])

            # Calculate age
            data['Вік'] = data['Дата народження'].apply(self.calculate_age)

            # Create a new Excel workbook and add sheets
            workbook = Workbook()
            all_sheet = workbook.active
            all_sheet.title = "all"
            younger_18_sheet = workbook.create_sheet(title="younger_18")
            age_18_45_sheet = workbook.create_sheet(title="18-45")
            age_45_70_sheet = workbook.create_sheet(title="45-70")
            older_70_sheet = workbook.create_sheet(title="older_70")

            # Write data to 'all' sheet
            for r_idx, row in enumerate(data.iterrows(), 1):
                for c_idx, value in enumerate(row[1], 1):
                    all_sheet.cell(row=r_idx, column=c_idx, value=value)

            # Write data to other sheets based on age categories
            for sheet, age_range in zip(
                    [younger_18_sheet, age_18_45_sheet, age_45_70_sheet, older_70_sheet],
                    [(0, 17), (18, 45), (46, 70), (71, 200)]
            ):
                age_filtered_data = data[data['Вік'].between(*age_range)]
                for r_idx, row in enumerate(age_filtered_data.iterrows(), 1):
                    for c_idx, value in enumerate(row[1], 1):
                        sheet.cell(row=r_idx, column=c_idx, value=value)

            # Save the Excel file
            workbook.save(self.xlsx_file_name)
            print("Ok")


        except pd.errors.ParserError:

            print(f"Error: Could not parse the CSV file: {self.csv_file_name}")

        except PermissionError:

            print(f"Error: Could not create the XLSX file: {self.xlsx_file_name} (Permission denied)")

        except Exception as e:

            print(f"Error: {e}")


# Usage:
xlsx_generator = EmployeeXLSXGenerator(
    csv_file_name="employees.csv",
    xlsx_file_name="employees.xlsx"
)
xlsx_generator.generate_xlsx()
